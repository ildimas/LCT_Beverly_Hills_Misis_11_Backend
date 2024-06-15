import sber_algo
import csv
from openpyxl import Workbook
from datetime import datetime

# пишем csv
filename_csv = "sample.csv"
headers = ["Компания", "Год счета", "Номер счета", "Позиция счета", 
           "Номер позиции распределения", "Дата отражения в учетной системе", "ID договора",
           "Услуга", "ID услуги", "Здание", "Класс ОС", "ID основного средства",
           "Признак использования в основной деятельности", "Признак передачи в аренду",
           "Площадь", "Сумма распределения", "Счет главной книги"]

with open(filename_csv, mode='w', newline='') as file:
    writer = csv.writer(file, delimiter=';')
    
    writer.writerow(headers)
    
    for row in sber_algo.y:
        writer.writerow(row)

print("CSV файл был успешно записан!")

# пишем xlsx
filename_xlsx = "sample.XLSX"
wb = Workbook()
ws = wb.active

for i in range(0, len(headers)):
    ws.cell(row = 1, column = i + 1).value = headers[i]

for i in range(0, len(sber_algo.y)):
    for j in range(0, len(sber_algo.y[i])):
        ws.cell(row = i + 2, column = j + 1).value = sber_algo.y[i][j]

for i in range(2, ws.max_row + 1):
    date_string = ws.cell(row = i, column = 6).value
    date_obj = datetime.strptime(date_string, "%Y-%m-%d")
    ws.cell(row = i, column = 6).value = date_obj
    ws.cell(row = i, column = 6).number_format = "DD.MM.YYYY"

wb.save(filename_xlsx)

print("XLSX файл был успешно записан!")